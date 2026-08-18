import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    {"key": "start", "label": "Startdatum & Tijd *", "example": "06.09.2026, 09:00", "width": 20},
    {"key": "end", "label": "Einddatum & Tijd *", "example": "06.09.2026, 15:00", "width": 20},
    {"key": "location.fromRoadNumber", "label": "Van Wegnummer *", "example": "A15", "width": 15},
    {"key": "location.fromRoadSide", "label": "Van Wegzijde *", "example": "Re", "width": 15},
    {"key": "location.fromMeter", "label": "Van km *", "example": "150,000", "width": 15},
    {"key": "location.betweenName", "label": "Van Plaats *", "example": "Gorinchem", "width": 20},
    {"key": "location.secondaryName", "label": "Tussen *", "example": "Leigraaf", "width": 20},
    {"key": "location.toRoadNumber", "label": "Naar Wegnummer *", "example": "A15", "width": 15},
    {"key": "location.toRoadSide", "label": "Naar Wegzijde *", "example": "Re", "width": 15},
    {"key": "location.toMeter", "label": "Tot km *", "example": "165,000", "width": 15},
    {"key": "location.andName", "label": "Naar Plaats *", "example": "Nijmegen", "width": 20},
    {"key": "location.primaryName", "label": "En *", "example": "Bemmel", "width": 20},
    {"key": "roadworkType", "label": "Wegwerktype", "example": "inspectie algemeen", "width": 22},
    {"key": "bestekId", "label": "Besteknummer", "example": "NL-31154600-inspecties voor MJPV", "width": 35},
    {"key": "managingDistrict", "label": "Wegbeheerder", "example": "ON District Zuid", "width": 22},
    {"key": "trafficHindranceClass", "label": "Hinderklasse", "example": "1 (geen file)", "width": 20},
    {"key": "roadblockType", "label": "Afzetsysteem", "example": "96a-430", "width": 18},
    {"key": "widthConstraint", "label": "Vrij doorrijdprofiel Breedte (m)", "example": "7,00", "width": 28},
    {"key": "isDraft", "label": "Opslaan als Concept? (Ja/Nee)", "example": "Ja", "width": 25}
]

def generate_excel_template(output_path, profile=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPIN Batch Aanvragen"

    # Styling header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write Headers
    for col_num, col_info in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=col_info["label"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col_num)].width = col_info["width"]

    # Write Example Row
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    example_font = Font(name="Calibri", size=10, italic=True, color="595959")
    
    for col_num, col_info in enumerate(COLUMNS, 1):
        val = col_info["example"]
        if profile:
            if col_info["key"] == "roadworkType" and profile.get("roadworkType"):
                val = profile["roadworkType"]
            elif col_info["key"] == "bestekId" and profile.get("bestekId"):
                val = profile["bestekId"]
            elif col_info["key"] == "managingDistrict" and profile.get("managingDistrict"):
                val = profile["managingDistrict"]
            elif col_info["key"] == "roadblockType" and profile.get("roadblockType"):
                val = profile["roadblockType"]

        cell = ws.cell(row=2, column=col_num, value=val)
        cell.fill = example_fill
        cell.font = example_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(output_path)
    return output_path

def parse_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Read header row mapping
    headers = []
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or "").strip()
        headers.append(val)

    key_map = {}
    for col_idx, h in enumerate(headers):
        for col_info in COLUMNS:
            if col_info["label"].lower() in h.lower() or col_info["key"].lower() in h.lower():
                key_map[col_idx] = col_info["key"]
                break

    measures = []
    for row in range(3, ws.max_row + 1):  # Skip row 1 (header) & row 2 (example)
        row_data = {}
        has_val = False
        for col_idx, key in key_map.items():
            cell_val = ws.cell(row=row, column=col_idx + 1).value
            if cell_val is not None:
                val_str = str(cell_val).strip()
                if val_str:
                    has_val = True
                    row_data[key] = val_str
        
        if has_val and row_data.get("start") and row_data.get("location.fromRoadNumber"):
            measures.append(row_data)

    return measures
